from random import triangular
from debugpy import trace_this_thread
from numpy import NaN
import openpyxl
import os

from sklearn.utils import column_or_1d

ROOT_DIR = os.path.dirname(os.path.abspath("curriculum - Copy - Copy - Copy.xlsx")) #root directory
REL_PATH = 'textMining/curriculum - Copy - Copy - Copy.xlsx' # relative path to root directory
COLUMN_COURSE = 'F' # column where course desc id present
COLUMN_WORDS = 'H' # column where unique words to be present
COLUMN_COUNT_CREATE = 'I' # column where count of unique create words to be present
COLUMN_COUNT_EVALUATE = 'J' # column where count of unique evaluate words to be present
COLUMN_COUNT_ANALYZE = 'K' # column where count of unique analyze words to be present
COLUMN_COUNT_APPLY = 'L' # column where count of unique apply words to be present
COLUMN_COUNT_UNDERSTAND = 'M' # column where count of unique understand words to be present
COLUMN_COUNT_REMEMBER = 'N' # column where count of unique remember words to be present
COLUMN_COUNT = 'O' # column where count of unique words to be present
COLUMN_COUNT_OF_WORDS = 'P' # column where count/mention of words to be present
COLUMN_UNIQUENESS = 'Q' # column where uniqueness of words to be present


create = {}
evaluate = {}
analyze = {}
apply = {}
understand = {}
remember = {}
mainDict = {}

print('-----------------')
print(ROOT_DIR)

dict = {'create':['create','assemble','construct','conjecture','develop','formulate','author','investigate','creates','created','creating','creation','generate','produce','design','make','fabricate','manufacture','build','erect','do','turn out','originate','invent','engender','devise','frame','shape','form','mould','forge','concoct','hatch','prompt','promote','foster','inspire','establish','found','institute','constitute','inaugrate','launch','start','appoint','invest','name','nominate','designate','put up','set up','raise','contrive','fashion','compose','model','engineer','guess', 'speculation', 'surmise', 'fancy', 'notion', 'belief', 'suspicion', 'presumption', 'assumption', 'theory', 'hypothesis', 'postulation', 'supposition','inference', 'extrapolation', 'projection','approximation', 'estimate', 'rough calculation', 'rough idea','guesswork', 'guessing', 'surmising', 'imagining', 'theorizing','speculate','infer','imagine','believe','think','suspect','presume','hypothesize','theorize','grow', 'evolve', 'mature', 'expand', 'enlarge', 'spread', 'advance', 'progress', 'prosper','succeed','thrive','flourish','blossom','bloom','burgeon','undertake','initiate','instigate','form','establish','supplement','reinforce','augment','extend','broaden','fill out','embellish','enhance','amplify','refine','improve','polish','perfect','unfold','ensue','contract','draw up','put together','coin','draft','express','articulate','utter','verbalize','phrase','render' ,'couch','voice','probe','scrutinize','inspect'],
        'evaluate':['evaluate','appraise','argue','defend','judge','select','support','value','critique','weigh','evaluated','evaluating','evaluator','assess','guage','estimate','rate','review','consider','price','quote','contend','assert','declare','maintain','proclaim','insist','held','aver','avow','reason','attest','expostulate','testify','swear','certify','propound','submit','posit','adduce','move','advocate','opine','allege','defend','explain','vindicate','justify','depose','asseverate','quarrel','disagree','row','squabble','bicker','wrangle','dispute','feud','dissent','clash','differ ','scrap','argufy','spat','altercate','debate','controvert','protect','guard','safeguard','preserve','secure','shield','shelter','screen','fortify','garrison','barricade','uphold','palliate','exonerate','explain','excuse','support','champion','endorse','sustain','bolster','justice','magistrate','conclude','decide','determine','consider','believe','think','deem ','view','deduce','gather','infer','tell','see','estimate','conjecture','try','hear','adjudicate','adjudge','decree','rule','find','arbitrate','mediate','moderate','evaluate','criticize','choose ','designate','name','nominate','appoint','elect','specify','stipulate','prefer','favour','bear','carry','prop up','keep up','bolster','brace','shore up','underpin','buttress','reinforce','confort','hearten','fortify','console','solace','reassure','succour','soothe','substantiate','corroborate','confirm','attest to','verify','prove','ratify','underwrite','subsidize','tolerate','abide','shoulder','assessment','appreciation','treatise','discourse','exposition','exegesis','oppress','burder','trouble','worry','beset','bother','upset','depress','distress','grieve','haunt','nag','torment','afflisct','perturb','plague'],
        'analyze':['analyze','differentiate','organize','relate','compare','contrast','distinguish','examine','experiment','question','test','analyzed','analyzes','analyzing','resolve','separate','reduce','decompose','disintegrate','dissect','divide','assay','test','examine','inspect','survey','scan','study','scrutinize','persue','search','explore','probe','research','check','sift','distinguish','discriminate','distinction','discern','recognize','identify','determine','characterize','individualize','individuate','transfrom','metamorphose','evolve','convert','change','modify','alter','adapt','order','arrange','assemble','marshal','group','dispose','classify','collocate','categorize','catalogue','codify','tabulate','compile','systematize','systemize','regulate','regiment','standardize','structure','shape','mould','pigeonhole','coordinate','orchestrate','choreograph','direct','run','manage','conduct','administrate','mobilize','mastermind','engineer','institute','form','create','establish','found','originate','begin','start','schedule','programme','timetable','tell','recount','narrate','describe','portray','depict','paint','unfold','set forth ','present','report','chronicle','outline','delineate','retail','recite','repeat','rehearse','relay','convey','communicate','impart','spin','detail','enumerate','list','specify','itemize','cite','particularize','connect','associate','correlate','ally','couple','bracket','bring together','relevant','relevance','concern','refer','reference','pertain','pertinent','bear','affect','involve','cover','touch','rapport','respond','understand','empathize','tune','collate','differentiate','balance','analogy','comparable','compete','match','resemble','emulate','rival','approach','nudge','difference','dissimilarity','dissimilitude','disparity','contradistinction','divergence','variance','variation','differentiation','incongruity','polarity','contradiction','opposite','antithesis','foil','perceive','observe','notice','spot','glimpse','detect','identify','inspect','inquire','sift','delve','probe','conside','vet','test','quiz','interrogate','cross-examine','cross-question','catechize','probe','investigation','trial','demonstration','examination','observation','query','quizzing','dubiousness','doubt','dispute','argument','controversy','reservation','issue','matter','problem','business','theme','subject','topic','proposal','contention','item','case','interview','appraisal','scrutinization','exploration','screening','criterion','proof','indication','measure','tax','strain'],
        'apply':['apply','execute','implement','solve','use','demonstrate','interpret','operate','schedule','sketch','applied','applying','try','bid','appeal','petition','entreaty','sue','register','audition','request','seek','solicit','claim','ask','appertain','bearing','pertain','pertinent','apposite','appropriate','fitting','germane','invlove','cover','spread','smear','exert','administer','exercise','employ','utilize','execute','prosecute','enact','accomplish','perform','implement','effect','archive','complete','enforce','discharge','prosecute','attain','fulfil','render','perform','enact','contrive','achieve','resolve','fathom','decipher','decode','break','interpret','translate','straighten','unravel','untangle','unfold','explain','expound','elucidate','utilize','avail','employ','work','operate','wield','ply','manoeuvre','manipulate','resort','exercise','employ','exert','practise','manage','handle','treat','behave','conduct','exploit','manipulate','capitalize','milk','trade on','misuse','mistreat','maltreat','trifle','consume','exhaust','deplete','spend','waste','fritter','squander','dissipate','usage','utilization','application','operation','manoeuvring','usefulness','advantage','exploitation','maltreatment','mistreatment','utility','service','gain','good','worth','motive','goal','sense','reason','necessity','demand','call','cause','grounds','justification','requirement','show','indicate','determine','establish','prove','validate','prove','confirm','verify','corroborate','substantiate','constitute','exibit','display','exemplify','illustrate','reveal','bespeak','signify','denote','manifest','testify','explain','elucidate','construe','intelligible','comprehend','transliterate','rewrite','steer','drive','pilot','efficacious','govern','superintend','plan','agenda','drawing','delineate','outline','recapitulate'],
        'understand':['understand','classify','describe','discuss','explain','identify','locate','recognize','report','select','translate','understanding','understood','comprehend','apprehend','grasp','see','perceive','discern','recognize','follow','fathom','penetrate','unravel','deciphetr','divine','master','envisage','appreciate','recognize','accept','sympathize','emphatize','accept','believe','infer','deduce','assume','surmise','gather','informed','discover','categorize','class','group','grade','rank','order','rate','type','codify','bracket','systemize','stratify','tabulate','list','file','assign','allocate','consign','label','report','narrate','recount','relate','express','represent','evoke','conjure','explain','expound','designate','pronounce','dub','characterize','hail','paint','outline','trace','converse','debate','confer','deliberate','ventilate','dispute','moot','explore','study','concern','consider','tackle','elucidate','expand','delineate','explicate','clarify','unfold','gloss','interpret','decipher','simplify','translate','illustrate','rationale','rationalize','legitimzie','mitigate','defend','recognize','point out','locate','discover','distinguish','recall','recollect','ascertain','discern','distinguish','confirm','verify','associate','link','connect','couple','bracket','commune','sympathize','equate','identical','pinpoint','unearth','reveal','ferret','uncover','stunble','chance','position', 'place','base','build','establish','put','found','station','fix','install','lodge','settle','seat','acknowlwdge','accept','concede','grant','confess','own','conscious','cognizant','apprehend','appreciate','certify','accredit','endorse','sanction','validate','uphold','support','commend','salute','applaud','reward','honour','homage','announce','describe','relay','divulge','publish','circulate','broadcast','blazon','declare','publicize','promulgate','document','chronicle','investigate','inquire','survey','research','study','tattle','accuse','arrive','appear','review','report','interpret','render','express','convert','transcribe','transliterate','elucidate','expound','clarify','unravel','convert','transform','alter','transplant'],
        'remember':['remember','define','duplicate','list','memorize','repeat','state','remembering','remembered','recall','recollect','remisce','muse','summon','retain','memorize','mindful','sure of','certain','commemorate','memorialize','bestow','explain','expound','elucidate','describe','clariify','precisely','determine','establish','stipulate','settle','demarcate','mark out','fix','bound','delimit','circumscribe','outline','silhouette','trace','copy','photocopy','carbon','fascimile','mimeograph','reprint','replice','reproduction','reproduce','twin','double','clone','match','mate','fellow','counterpart','identical','matching','corresponding','equivalent','matched','paired','twofold','coupled','reproduce','inventory','catalogue','record','register','roll','file','index','directory','listing','checklist','tally','docket','enumeration','tabulation','series','enumerate','chronicle','categorize','itemize','classify','group','arrange','file','log','minute','rank','alphabetize','retain','restate','reiterate','rehearse','capitulate','quote','echo','redo','rebroadcast','rerun','reshow','replay']}

overallDict = {}


def titleToNumber(s):
        r = 0
        starting = len(s) - 1
        for char in s:
            r += (ord(char) - ord('A') + 1) * 26 ** starting
            starting -= 1 
        return r


def readingFile_SpecificColumn(rel_path,column):
    abs_file_path = os.path.join(ROOT_DIR, rel_path)
    book = openpyxl.load_workbook(abs_file_path)

    sheet = book.active 
    print('max rows = ',sheet.max_row)
    print('max columns = ',sheet.max_column)

    for cell in sheet[column]:
        for key,values in dict.items():
            if cell.value in dict.values():
                print(key,'',values)

    
    overallCreate = 0
    overallEvaluate = 0
    overallAnalyze = 0
    overallApply = 0
    overallUnderstand = 0
    overallRemember = 0
    colNumber_words = titleToNumber(COLUMN_WORDS)
    colNumber_count = titleToNumber(COLUMN_COUNT)
    colNumber_uniqueness = titleToNumber(COLUMN_UNIQUENESS)
    colNumber_count_create = titleToNumber(COLUMN_COUNT_CREATE)
    colNumber_count_evaluate = titleToNumber(COLUMN_COUNT_EVALUATE)
    colNumber_count_analyze = titleToNumber(COLUMN_COUNT_ANALYZE)
    colNumber_count_remember = titleToNumber(COLUMN_COUNT_REMEMBER)
    colNumber_count_understand = titleToNumber(COLUMN_COUNT_UNDERSTAND)
    colNumber_count_apply = titleToNumber(COLUMN_COUNT_APPLY)
    colNumber_count_of_words = titleToNumber(COLUMN_COUNT_OF_WORDS)
    sheet.cell(row=1,column=colNumber_count).value = 'Total count of unique words'
    sheet.cell(row=1,column=colNumber_words).value = 'Unique words'
    sheet.cell(row=1,column=colNumber_count_analyze).value = 'Unique \'analyse\' words'
    sheet.cell(row=1,column=colNumber_count_apply).value = 'Unique \'apply\' words'
    sheet.cell(row=1,column=colNumber_count_create).value = 'Unique \'create\' words'
    sheet.cell(row=1,column=colNumber_count_evaluate).value = 'Unique \'evaluate\' words'
    sheet.cell(row=1,column=colNumber_count_remember).value = 'Unique \'remember\' words'
    sheet.cell(row=1,column=colNumber_count_understand).value = 'Unique \'understand\' words'
    
    for cell in sheet[column]:    
        print('---------------')
        cellRow = cell.row
        print(cellRow)
        newDict = {}
        if cell.value is not None:
            for key, value in dict.items():
                counterCreate = 0
                counterEvaluate = 0
                counterAnalyze = 0
                counterApply = 0
                counterUnderstand = 0
                counterRemember = 0

                create = {}
                evaluate = {}
                analyze = {}
                apply = {}
                understand = {}
                remember = {}

                for v in value:
                    if v in cell.value.lower():
                        print('row =',cell.row,'-',key,'=',v)
                        if key == 'create':
                            if sheet.cell(row=cell.row,column=colNumber_words).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_words).value = sheet.cell(row=cell.row,column=colNumber_words).value + ',' + v
                            else:
                                sheet.cell(row=cell.row,column=colNumber_words).value = v
                            if sheet.cell(row=cell.row,column=colNumber_count).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count).value = sheet.cell(row=cell.row,column=colNumber_count).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count).value = 1
                            counterCreate += 1
                            overallCreate += 1
                            if sheet.cell(row=cell.row,column=colNumber_count_create).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count_create).value = sheet.cell(row=cell.row,column=colNumber_count_create).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count_create).value = counterCreate
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            book.save(abs_file_path)
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                        elif key == 'evaluate':
                            if sheet.cell(row=cell.row,column=colNumber_words).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_words).value = sheet.cell(row=cell.row,column=colNumber_words).value + ',' + v
                            else:
                                sheet.cell(row=cell.row,column=colNumber_words).value = v
                            if sheet.cell(row=cell.row,column=colNumber_count).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count).value = sheet.cell(row=cell.row,column=colNumber_count).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count).value = 1
                            counterEvaluate += 1
                            overallEvaluate += 1
                            if sheet.cell(row=cell.row,column=colNumber_count_evaluate).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count_evaluate).value = sheet.cell(row=cell.row,column=colNumber_count_evaluate).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count_evaluate).value = counterEvaluate
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                        elif key == 'analyze':
                            if sheet.cell(row=cell.row,column=colNumber_words).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_words).value = sheet.cell(row=cell.row,column=colNumber_words).value + ',' + v
                            else:
                                sheet.cell(row=cell.row,column=colNumber_words).value = v
                            if sheet.cell(row=cell.row,column=colNumber_count).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count).value = sheet.cell(row=cell.row,column=colNumber_count).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count).value = 1
                            counterAnalyze += 1
                            overallAnalyze += 1
                            if sheet.cell(row=cell.row,column=colNumber_count_analyze).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count_analyze).value = sheet.cell(row=cell.row,column=colNumber_count_analyze).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count_analyze).value = counterAnalyze
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                        elif key == 'apply':
                            if sheet.cell(row=cell.row,column=colNumber_words).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_words).value = sheet.cell(row=cell.row,column=colNumber_words).value + ',' + v
                            else:
                                sheet.cell(row=cell.row,column=colNumber_words).value = v
                            if sheet.cell(row=cell.row,column=colNumber_count).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count).value = sheet.cell(row=cell.row,column=colNumber_count).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count).value = 1
                            counterApply += 1
                            overallApply += 1
                            if sheet.cell(row=cell.row,column=colNumber_count_apply).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count_apply).value = sheet.cell(row=cell.row,column=colNumber_count_apply).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count_apply).value = counterApply
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                        elif key == 'understand':
                            if sheet.cell(row=cell.row,column=colNumber_words).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_words).value = sheet.cell(row=cell.row,column=colNumber_words).value + ',' + v
                            else:
                                sheet.cell(row=cell.row,column=colNumber_words).value = v
                            if sheet.cell(row=cell.row,column=colNumber_count).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count).value = sheet.cell(row=cell.row,column=colNumber_count).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count).value = 1
                            counterUnderstand += 1
                            overallUnderstand += 1
                            if sheet.cell(row=cell.row,column=colNumber_count_understand).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count_understand).value = sheet.cell(row=cell.row,column=colNumber_count_understand).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count_understand).value = counterUnderstand
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                        elif key == 'remember':
                            if sheet.cell(row=cell.row,column=colNumber_words).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_words).value = sheet.cell(row=cell.row,column=colNumber_words).value + ',' + v
                            else:
                                sheet.cell(row=cell.row,column=colNumber_words).value = v
                            if sheet.cell(row=cell.row,column=colNumber_count).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count).value = sheet.cell(row=cell.row,column=colNumber_count).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count).value = 1
                            counterRemember += 1
                            overallRemember += 1
                            if sheet.cell(row=cell.row,column=colNumber_count_remember).value is not None:
                                sheet.cell(row=cell.row,column=colNumber_count_remember).value = sheet.cell(row=cell.row,column=colNumber_count_remember).value + 1
                            else:
                                sheet.cell(row=cell.row,column=colNumber_count_remember).value = counterRemember
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
        # if sheet.cell(row=cell.row,column=16).value is not None:
        #     sheet.cell(row=cell.row,column=16).value = create
        # else:
        mainDict[str(cellRow)] = newDict
        #print('create-',create,'\n','apply-',apply,'\n','analyze-',analyze,'\n','remember-',remember,'\n','understand-',understand,'\n','evaluate-',evaluate)

    print(mainDict)
    overallUniqueWords = overallCreate+overallAnalyze+overallApply+overallEvaluate+overallRemember+overallUnderstand
    print(overallDict)



    #uniqueness for overall count
    for cell in sheet[column]:
        if cell.value is not None:
            print(cell.row,' - Calculating uniqueness of the words')
            if cell.row == 1:
                sheet.cell(row=cell.row,column=colNumber_uniqueness).value = 'Uniqueness of words'
                sheet.cell(row=cell.row,column=colNumber_count_of_words).value = 'Count/mention of the words in this description'
                book.save(abs_file_path)
            else:
                # print(str(mainDict[str(cell.row)]))
                sheet.cell(row=cell.row,column=colNumber_count_of_words).value = str(mainDict[str(cell.row)]).replace('{','').replace('}','')
                for keys in overallDict:
                    if(keys in cell.value.lower()):
                        if sheet.cell(row=cell.row,column=colNumber_uniqueness).value is not None:
                            sheet.cell(row=cell.row,column=colNumber_uniqueness).value = sheet.cell(row=cell.row,column=colNumber_uniqueness).value + '\n' + keys + ':' + str(((overallDict[keys]/overallUniqueWords)*100)) + '%'
                            book.save(abs_file_path)
                        else:
                            sheet.cell(row=cell.row,column=colNumber_uniqueness).value = keys + ':' + str(((overallDict[keys]/overallUniqueWords)*100)) + '%'
                            book.save(abs_file_path)

print('file operations completed')
            

####################################################################33

if __name__ == "__main__":
    sheet = readingFile_SpecificColumn(rel_path = REL_PATH,column=COLUMN_COURSE)