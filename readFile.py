from numpy import NaN
import openpyxl
import os

from sklearn.utils import column_or_1d

ROOT_DIR = os.path.dirname(os.path.abspath("newFile.xlsx"))
REL_PATH = 'textMining/newFile.xlsx'
COLUMN_COURSE = 'F'

print('-----------------')
print(ROOT_DIR)

dict = {'create':['design','Introduction','survey','emphasis','study','assemble','construct','conjecture','develop','formulate','author','investigate'],
        'evaluate':['appraise','argue','defend','judge','select','support','value','critique','weigh'],
        'analyze':['differentiate','organize','relate','compare','contrast','distinguish','examine','experiment','question','test'],
        'apply':['execute','implement','solve','use','demonstrate','interpret','operate','schedule','sketch'],
        'understand':['classify','describe','discuss','explain','identify','locate','recognize','report','select','translate'],
        'remember':['define','duplicate','list','memorize','repeat','state','probation']}

create = {}
evaluate = {}
analyze = {}
apply = {}
understand = {}
remember = {}
mainDict = {}
overallDict = {}
overallUniqueWords = 0

def fillColumnWords(sheet, rowNumber, colNumber, value):
    colNumber = 8
    print(sheet.cell(row=rowNumber,column=colNumber).value)
    if sheet.cell(row=rowNumber,column=colNumber).value is not None:
        sheet.cell(row=rowNumber,column=colNumber).value = sheet.cell(row=rowNumber,column=colNumber).value + ',' + value
    else:
        sheet.cell(row=rowNumber,column=colNumber).value = sheet.cell(row=rowNumber,column=colNumber).value
    return sheet

def fillColumnNumbers(sheet, rowNumber, colNumber):
    colNumber = 20
    print(sheet.cell(row=rowNumber,column=colNumber).value)
    if sheet.cell(row=rowNumber,column=colNumber).value is not None:
        sheet.cell(row=rowNumber,column=colNumber).value = sheet.cell(row=rowNumber,column=colNumber).value + 1
    else:
        sheet.cell(row=rowNumber,column=colNumber).value = 1
    return sheet

def readingFile_SpecificColumn(rel_path,column):
    # rel_path = "textMining/curriculum.xlsx"
    abs_file_path = os.path.join(ROOT_DIR, rel_path)
    book = openpyxl.load_workbook(abs_file_path)

    sheet = book.active 
    print(sheet.max_row)
    print(sheet.max_column)

    # cells = sheet['A2':str(chr(sheet.max_column+65))+str(sheet.max_row)]
    # for value in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column,values_only=True):
    #         print(value)

    # for cell in sheet[column]:
    #         print(f"{cell.value}")


    for cell in sheet[column]:
        for key,values in dict.items():
            if cell.value in dict.values():
                print(key,'',values)

    counterCreate = 0
    counterEvaluate = 0
    counterAnalyze = 0
    counterApply = 0
    counterUnderstand = 0
    counterRemember = 0
    overallCreate = 0
    overallEvaluate = 0
    overallAnalyze = 0
    overallApply = 0
    overallUnderstand = 0
    overallRemember = 0
    
    for cell in sheet[column]:    
        # for key in dict:
        #     for value in values[key]:
        #         if value in cell.value:
        #             print(key)
        # print(any(any(s in cell.value for s in sublist)for sublist in dict.values()))
        # print(key for key,value in dict.items() if any(s in cell.value for s in value))
                        # print(f"{cell.value}")
        print('---------------')
        counter=0
        cellRow = cell.row
        print(cellRow)
        newDict = {}
        for key, value in dict.items():
            counterCreate = 0
            counterEvaluate = 0
            counterAnalyze = 0
            counterApply = 0
            counterUnderstand = 0
            counterRemember = 0

            create = {}
            evaluate = {}
            analyze = {}
            apply = {}
            understand = {}
            remember = {}

            for v in value:
                if v in cell.value:
                    print(cell.row,'-',key,'-',v)
                    if key == 'create':
                        print(sheet.cell(row=cell.row,column=8).value)
                        if sheet.cell(row=cell.row,column=8).value is not None:
                            sheet.cell(row=cell.row,column=8).value = sheet.cell(row=cell.row,column=8).value + ',' + v
                        else:
                            sheet.cell(row=cell.row,column=8).value = v
                        print(sheet.cell(row=cell.row,column=20).value)
                        if sheet.cell(row=cell.row,column=20).value is not None:
                            sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                        else:
                            sheet.cell(row=cell.row,column=20).value = 1
                        # if sheet.cell(row=cell.row,column=8).value is not None:
                            # sheet = fillColumnWords(sheet,cell.row,column,v)
                            #sheet.cell(row=cell.row,column=8).value = sheet.cell(row=cell.row,column=8).value + ',' + v
                        counterCreate += 1
                        overallCreate += 1
                            # sheet.cell(row=cell.row,column=14).value = counterCreate
                        book.save(abs_file_path)
                            # sheet = fillColumnNumbers(sheet, cell.row, column)
                            # if sheet.cell(row=cell.row,column=20).value is not None:
                            #     sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                            # else:
                            #     sheet.cell(row=cell.row,column=20).value = counterCreate
                            # book.save(abs_file_path)
                            # if v not in newDict:
                            #     newDict[v] = 1
                            # else:
                            #     newDict += 1
                        # overallUniqueWords += 1
                        if v not in overallDict:
                            overallDict[v] = 1
                        else:
                            overallDict[v] += 1
                        print('ok')
                        # else:
                        #     sheet.cell(row=cell.row,column=8).value = v
                        #     counterCreate += 1
                        #     overallCreate += 1
                        #     sheet.cell(row=cell.row,column=14).value = counterCreate
                        #     if sheet.cell(row=cell.row,column=20).value is not None:
                        #         sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                        #     else:
                        #         sheet.cell(row=cell.row,column=20).value = counterCreate
                        #     book.save(abs_file_path)
                        #     if v not in newDict:
                        #         newDict[v] = 1
                        #     else:
                        #         newDict[v] += 1
                        #     if v not in overallDict:
                        #         overallDict[v] = 1
                        #     else:
                        #         overallDict[v] += 1
                        #     print('ok')
                    elif key == 'evaluate':
                        if sheet.cell(row=cell.row,column=9).value is not None:
                            sheet.cell(row=cell.row,column=9).value = sheet.cell(row=cell.row,column=9).value + ',' + v
                            counterEvaluate += 1
                            overallEvaluate += 1
                            sheet.cell(row=cell.row,column=15).value = counterEvaluate
                            if sheet.cell(row=cell.row,column=20).value is not None:
                                sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                            else:
                                sheet.cell(row=cell.row,column=20).value = counterEvaluate
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                        else:
                            sheet.cell(row=cell.row,column=9).value = v
                            counterEvaluate += 1
                            overallEvaluate += 1
                            sheet.cell(row=cell.row,column=15).value = counterEvaluate
                            if sheet.cell(row=cell.row,column=20).value is not None:
                                sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                            else:
                                sheet.cell(row=cell.row,column=20).value = counterEvaluate
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                    elif key == 'analyze':
                        if sheet.cell(row=cell.row,column=10).value is not None:
                            sheet.cell(row=cell.row,column=10).value = sheet.cell(row=cell.row,column=10).value + ',' + v
                            counterAnalyze += 1
                            overallAnalyze += 1
                            sheet.cell(row=cell.row,column=16).value = counterAnalyze
                            if sheet.cell(row=cell.row,column=20).value is not None:
                                sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                            else:
                                sheet.cell(row=cell.row,column=20).value = counterAnalyze
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                        else:
                            sheet.cell(row=cell.row,column=10).value = v
                            counterAnalyze += 1
                            overallAnalyze += 1
                            sheet.cell(row=cell.row,column=16).value = counterAnalyze
                            if sheet.cell(row=cell.row,column=20).value is not None:
                                sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                            else:
                                sheet.cell(row=cell.row,column=20).value = counterAnalyze
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                    elif key == 'apply':
                        if sheet.cell(row=cell.row,column=11).value is not None:
                            sheet.cell(row=cell.row,column=11).value = sheet.cell(row=cell.row,column=11).value + ',' + v
                            counterApply += 1
                            overallApply += 1
                            sheet.cell(row=cell.row,column=17).value = counterApply
                            if sheet.cell(row=cell.row,column=20).value is not None:
                                sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                            else:
                                sheet.cell(row=cell.row,column=20).value = counterApply
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                        else:
                            sheet.cell(row=cell.row,column=11).value = v
                            counterApply += 1
                            overallApply += 1
                            sheet.cell(row=cell.row,column=17).value = counterApply
                            if sheet.cell(row=cell.row,column=20).value is not None:
                                sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                            else:
                                sheet.cell(row=cell.row,column=20).value = counterApply
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                    elif key == 'understand':
                        if sheet.cell(row=cell.row,column=12).value is not None:
                            sheet.cell(row=cell.row,column=12).value = sheet.cell(row=cell.row,column=12).value + ',' + v
                            counterUnderstand += 1
                            overallUnderstand += 1
                            sheet.cell(row=cell.row,column=18).value = counterUnderstand
                            if sheet.cell(row=cell.row,column=20).value is not None:
                                sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                            else:
                                sheet.cell(row=cell.row,column=20).value = counterUnderstand
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                        else:
                            sheet.cell(row=cell.row,column=12).value = v
                            counterUnderstand += 1
                            overallUnderstand += 1
                            sheet.cell(row=cell.row,column=18).value = counterUnderstand
                            if sheet.cell(row=cell.row,column=20).value is not None:
                                sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                            else:
                                sheet.cell(row=cell.row,column=20).value = counterUnderstand
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                    elif key == 'remember':
                        if sheet.cell(row=cell.row,column=13).value is not None:
                            sheet.cell(row=cell.row,column=13).value = sheet.cell(row=cell.row,column=13).value + ',' + v
                            counterRemember += 1
                            overallRemember += 1
                            sheet.cell(row=cell.row,column=19).value = counterRemember
                            if sheet.cell(row=cell.row,column=20).value is not None:
                                sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                            else:
                                sheet.cell(row=cell.row,column=20).value = counterRemember
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                        else:
                            sheet.cell(row=cell.row,column=13).value = v
                            counterRemember += 1
                            overallRemember += 1
                            sheet.cell(row=cell.row,column=19).value = counterRemember
                            if sheet.cell(row=cell.row,column=20).value is not None:
                                sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + 1
                            else:
                                sheet.cell(row=cell.row,column=20).value = counterRemember
                            book.save(abs_file_path)
                            if v not in newDict:
                                newDict[v] = 1
                            else:
                                newDict[v] += 1
                            if v not in overallDict:
                                overallDict[v] = 1
                            else:
                                overallDict[v] += 1
                            print('ok')
                    
        # print the % of each word in the total unique words in each row
        print('===========create=============')
        for key,value in create.items():
            print(key,'-',value)
        print('===========evaluate=============')
        for key,value in evaluate.items():
            print(key,'-',value)
        print('===========analyze=============')
        for key,value in analyze.items():
            print(key,'-',value)
        print('===========apply=============')
        for key,value in apply.items():
            print(key,'-',value)
        print('===========understand=============')
        for key,value in understand.items():
            print(key,'-',value)
        print('===========remember=============')
        for key,value in remember.items():
            print(key,'-',value)

        mainDict[str(cellRow)] = newDict

    print(overallCreate,overallEvaluate,overallAnalyze,overallApply,overallUnderstand,overallRemember)
    overallUniqueWords = overallCreate+overallAnalyze+overallApply+overallEvaluate+overallRemember+overallUnderstand

    print(mainDict)
    print(overallDict)




    #uniqueness for each row wise
    # for cell in sheet[column]:
    #     for rowNumber in mainDict:
    #         print('->>',rowNumber,'-->',cell.row)
    #         print(mainDict[rowNumber])
    #         if rowNumber != '1':
    #             if int(rowNumber) == cell.row:
    #                 currentDict = mainDict[rowNumber]
    #                 for keyWord in currentDict:
    #                     print(rowNumber,keyWord,currentDict[keyWord])
    #                     if sheet.cell(row=cell.row,column=26).value is not None:
    #                         sheet.cell(row=cell.row,column=26).value = sheet.cell(row=cell.row,column=26).value + '\n' + keyWord + ':' + str(((currentDict[keyWord]/overallUniqueWords)*100))
    #                         book.save(abs_file_path)
    #                     else:
    #                         sheet.cell(row=cell.row,column=26).value = keyWord + ':' + str(((currentDict[keyWord]/overallUniqueWords)*100))
    #                         book.save(abs_file_path)
    #             else:
    #                 print('ok 1sst row')
    #         else:
    #             print('ok row 1')



    #uniqueness for overall count
    for cell in sheet[column]:
        for keys in overallDict:
            if(keys in cell.value):
                if sheet.cell(row=cell.row,column=26).value is not None:
                    sheet.cell(row=cell.row,column=26).value = sheet.cell(row=cell.row,column=26).value + '\n' + keys + ':' + str(((overallDict[keys]/overallUniqueWords)*100))
                    book.save(abs_file_path)
                else:
                    sheet.cell(row=cell.row,column=26).value = keys + ':' + str(((overallDict[keys]/overallUniqueWords)*100))
                    book.save(abs_file_path)
            


    # for cell in sheet[column]:
    #     for key in create:
    #             if key in cell.value:
    #                 print(key)
    #                 if sheet.cell(row=cell.row,column=20).value is not None:
    #                     sheet.cell(row=cell.row,column=20).value = sheet.cell(row=cell.row,column=20).value + ',' + v + '-' + create[v]/overallCreate+overallApply+overallAnalyze+overallEvaluate+overallRemember+overallUnderstand
    #                 else:
    #                     sheet.cell(row=cell.row,column=20).value = v + '-' + create[v]/overallCreate+overallApply+overallAnalyze+overallEvaluate+overallRemember+overallUnderstand
    
    # return sheet

####################################################################33

def classifyWordsInCategories(dict,sheet,column):
    for cell in sheet[column]:
        for key,values in dict.items():
            if cell.value in dict.values():
                print(key,'',values)
    counter=0
    for cell in sheet[column]:    
        # for key in dict:
        #     for value in values[key]:
        #         if value in cell.value:
        #             print(key)
        # print(any(any(s in cell.value for s in sublist)for sublist in dict.values()))
        # print(key for key,value in dict.items() if any(s in cell.value for s in value))
                        # print(f"{cell.value}")
        print('---------------')
        counter=0
        for key, value in dict.items():
            for v in value:
                if v in cell.value:
                    print(cell.row,'-',key,'-',v)
                    if key == 'create':
                        sheet.cell(row=cell.row,column=8).value = v
                        print('ok')
                    elif key == 'evaluate':
                        sheet.cell(row=cell.row,column=9).value = v
                        print('ok')
                    elif key == 'analyse':
                        sheet.cell(row=cell.row,column=10).value = v
                        print('ok')
                    elif key == 'apply':
                        sheet.cell(row=cell.row,column=11).value = v
                        print('ok')
                    elif key == 'understand':
                        sheet.cell(row=cell.row,column=12).value = v
                        print('ok')
                    elif key == 'remember':
                        sheet.cell(row=cell.row,column=13).value = v
                        print('ok')


if __name__ == "__main__":
    sheet = readingFile_SpecificColumn(rel_path = REL_PATH,column=COLUMN_COURSE)
    # iterating_column("books.xlsx", sheet_name="Sheet 1 - Books",
    #                 col="A")
    #classifyWordsInCategories(dict,sheet,column = 'F')